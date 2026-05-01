"""
XML application form parser and exporter.

Parses Appform application export packages (view.xml) and generates:
- Excel tables (.xlsx)        requires: pandas, openpyxl
- CSV tables (.csv)           requires: pandas
- Markdown tables (.md)       always available
- SDK job_submit.yaml (.yaml) requires: PyYAML

Core XML parsing uses only the standard library.
"""

import glob
import os
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Lazy availability checks for optional dependencies
# ---------------------------------------------------------------------------


def _has_yaml():
    try:
        import yaml  # noqa: F401

        return True
    except ImportError:
        return False


def _has_pandas():
    try:
        import pandas  # noqa: F401

        return True
    except ImportError:
        return False


def _has_openpyxl():
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


# ---------------------------------------------------------------------------
# Core XML parsing (stdlib only)
# ---------------------------------------------------------------------------


def parse_dynamic_scripts(base_dir: str, app_id: str) -> Dict[str, Dict[str, str]]:
    """Parse dynamic script files to extract option values."""
    options_dict: Dict[str, Dict[str, str]] = {}
    script_dir = os.path.join(base_dir, "script")

    if not os.path.exists(script_dir):
        return options_dict

    dynamic_scripts = glob.glob(os.path.join(script_dir, "dynamic_*.sh"))

    for script_file in dynamic_scripts:
        try:
            with open(script_file, "r", encoding="utf-8") as f:
                content = f.read()

            pattern = r'(\w+)\s*=\s*["\']([^"\']+)["\']'
            matches = re.findall(pattern, content)

            for var_name, var_value in matches:
                if var_name.startswith("JH_") or var_name.startswith("STAR_"):
                    values = []
                    labels = []
                    parts = var_value.split(";")
                    for part in parts:
                        if "|" in part:
                            value, label = part.split("|", 1)
                            values.append(value.strip())
                            labels.append(label.strip())
                        else:
                            values.append(part.strip())
                            labels.append(part.strip())

                    options_dict[var_name] = {
                        "values": ";".join(values),
                        "labels": ";".join(labels),
                        "raw": var_value,
                    }
        except Exception:
            pass

    return options_dict


def parse_xml_to_table(
    xml_file: str, base_dir: Optional[str] = None
) -> Tuple[List[Dict[str, str]], str]:
    """
    Parse an XML file and extract form data.

    Returns:
        Tuple of (list_of_row_dicts, app_id_string)
    """
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()

        i18n_dict: Dict[str, str] = {}
        i18ns_section = root.find("i18ns")
        if i18ns_section is not None:
            for i18n in i18ns_section.findall("i18n"):
                key = i18n.get("key", "")
                zh = i18n.get("zh", "")
                i18n_dict[key] = zh

        form_data: List[Dict[str, str]] = []
        app_ids: set = set()

        dynamic_options: Dict[str, Dict[str, str]] = {}
        if base_dir:
            job_configs_section = root.find("jobSubmitConfigs")
            if job_configs_section is not None:
                first_config = job_configs_section.find("jobSubmitConfig")
                if first_config is not None:
                    app_id = first_config.get("appId", "")
                    dynamic_options = parse_dynamic_scripts(base_dir, app_id)

        job_configs_section = root.find("jobSubmitConfigs")
        if job_configs_section is not None:
            for config in job_configs_section.findall("jobSubmitConfig"):
                app_id = config.get("appId", "")
                app_ids.add(app_id)
                param_name = config.findtext("paramName", "")
                sign = config.findtext("sign", "")
                required = config.findtext("required", "false")
                validation = config.findtext("validation", "")
                default_value = config.findtext("defaultValue", "")

                i18n_key = f"{app_id}.{param_name}.label"
                chinese_desc = i18n_dict.get(i18n_key, "")

                option_values = ""
                option_labels = ""
                if param_name in dynamic_options:
                    option_values = dynamic_options[param_name]["values"]
                    option_labels = dynamic_options[param_name]["labels"]

                form_data.append(
                    {
                        "appid/应用 id": app_id,
                        "字段": param_name,
                        "字段中文说明": chinese_desc,
                        "默认值": default_value,
                        "字段类型": sign,
                        "是否必填": "TRUE" if required.lower() == "true" else "",
                        "正则过滤字符": validation,
                        "选项值": option_values,
                        "选项显示": option_labels,
                        "源文件": xml_file,
                    }
                )

        app_id_str = "_".join(sorted(app_ids)) if app_ids else "unknown"
        return form_data, app_id_str

    except Exception as e:
        print(f"Error parsing {xml_file}: {e}")
        return [], "error"


def get_base_dir(xml_file: str) -> str:
    """Get the application root directory for an XML file."""
    current_dir = os.path.dirname(os.path.abspath(xml_file))

    for _ in range(3):
        script_dir = os.path.join(current_dir, "script")
        if os.path.exists(script_dir):
            return current_dir
        parent_dir = os.path.dirname(current_dir)
        if parent_dir == current_dir:
            break
        current_dir = parent_dir

    return os.path.dirname(os.path.abspath(xml_file))


def find_xml_files(
    directory: str = ".", pattern: str = "view.xml", recursive: bool = True
) -> List[str]:
    """
    Find XML files in a directory.

    Args:
        directory: Root directory to search
        pattern: Filename pattern (default: view.xml)
        recursive: Search subdirectories

    Returns:
        Sorted list of XML file paths
    """
    xml_files: List[str] = []
    exclude_files = ["package.xml", "pom.xml", "web.xml", "settings.xml", "Appinfo.xml"]

    if recursive:
        for root, dirs, files in os.walk(directory):
            dirs[:] = [
                d
                for d in dirs
                if not d.startswith(".")
                and d not in ["node_modules", "__pycache__", "venv"]
            ]
            for file in files:
                if file.endswith(".xml") and file not in exclude_files:
                    if (
                        pattern == "*"
                        or pattern == "*.xml"
                        or file == pattern
                        or (pattern.endswith(".xml") and file.endswith(".xml"))
                        or re.match(pattern.replace("*", ".*"), file)
                    ):
                        xml_files.append(os.path.join(root, file))
    else:
        matches = glob.glob(os.path.join(directory, pattern))
        xml_files = [f for f in matches if os.path.basename(f) not in exclude_files]

    return sorted(xml_files)


def collect_xml_data(xml_files: List[str]) -> Tuple[List[Dict[str, str]], List[str]]:
    """
    Parse multiple XML files and collect all data.

    Returns:
        Tuple of (all_data, source_files)
    """
    all_data: List[Dict[str, str]] = []
    source_files: List[str] = []

    for xml_file in xml_files:
        base_dir = get_base_dir(xml_file)
        data, _ = parse_xml_to_table(xml_file, base_dir)
        if data:
            all_data.extend(data)
            source_files.append(xml_file)

    return all_data, source_files


# ---------------------------------------------------------------------------
# SDK YAML format (requires PyYAML)
# ---------------------------------------------------------------------------

_SIGN_TO_TYPE = {
    "text": "text",
    "input": "text",
    "select": "select",
    "switch": "switch",
    "upload": "upload",
    "file": "upload",
    "textarea": "text",
    "number": "text",
    "password": "text",
    "hidden": "text",
}

_KNOWN_CLI_ARGS = {
    "JH_CAS": {"cli_arg": "input", "short_arg": "i"},
    "JH_DAT": {"cli_arg": "dat", "short_arg": None},
    "JH_JOB_NAME": {"cli_arg": "job-name", "short_arg": None},
    "JH_JOB_CONF": {"cli_arg": "job-conf", "short_arg": None},
    "JH_NODE_GROUP": {"cli_arg": "group", "short_arg": "g"},
    "JH_NCPU": {"cli_arg": "ncpu", "short_arg": "n"},
    "JH_NCPU2": {"cli_arg": "ncpu2", "short_arg": "n2"},
    "JH_JAVA": {"cli_arg": "java", "short_arg": "j"},
    "JH_JAVA2": {"cli_arg": "java2", "short_arg": "j2"},
    "JH_RELEASE": {"cli_arg": "release", "short_arg": "r"},
    "JH_OTHERS": {"cli_arg": "others", "short_arg": None},
    "JH_JOB_WORK_PATH": {"cli_arg": "job-work-path", "short_arg": None},
    "STAR_POST_SWITCH": {"cli_arg": "post", "short_arg": "post"},
    "STAR_POST_NODEGROUP": {"cli_arg": "group2", "short_arg": "g2"},
    "JH_NNODE": {"cli_arg": "nnode", "short_arg": "N"},
    "JH_MEMORY": {"cli_arg": "memory", "short_arg": "m"},
    "JH_NGPU": {"cli_arg": "ngpu", "short_arg": None},
    "JH_NCPU_MESH": {"cli_arg": "ncpu-mesh", "short_arg": "nm"},
    "JH_PLATFORM": {"cli_arg": "platform", "short_arg": None},
    "JH_QUEUE": {"cli_arg": "queue", "short_arg": "q"},
    "SUB_QUEUE": {"cli_arg": "queue", "short_arg": "q"},
}


def create_sdk_yaml(all_data: List[Dict[str, str]], output_file: str) -> None:
    """
    Generate a job_submit.yaml compatible with Appform SDK / JobProfileManager.

    Requires: PyYAML (``pip install PyYAML``)
    """
    if not _has_yaml():
        raise ImportError(
            "PyYAML is required for YAML output. Install: pip install PyYAML"
        )

    import yaml

    if not all_data:
        print(f"  Warning: no data, skipping YAML: {output_file}")
        return

    apps: Dict[str, List[Dict[str, str]]] = {}
    for row in all_data:
        app_id = row["appid/应用 id"]
        if app_id not in apps:
            apps[app_id] = []
        apps[app_id].append(row)

    applications: Dict[str, Any] = {}

    for app_id, rows in apps.items():
        parameters: List[Dict[str, Any]] = []

        for row in rows:
            param_name = row["字段"]
            sign = row.get("字段类型", "text")
            required = row.get("是否必填", "") == "TRUE"
            validation = row.get("正则过滤字符", "") or None
            default_value = row.get("默认值", "") or None
            description = row.get("字段中文说明", "") or param_name

            param_type = _SIGN_TO_TYPE.get(sign.lower(), "text")

            param: Dict[str, Any] = {
                "name": param_name,
                "type": param_type,
                "required": required,
                "description": description,
            }

            if validation:
                param["validation"] = validation

            if default_value is not None and str(default_value).strip():
                param["default"] = str(default_value)

            if param_name in _KNOWN_CLI_ARGS:
                cli_info = _KNOWN_CLI_ARGS[param_name]
                if cli_info["cli_arg"]:
                    param["cli_arg"] = cli_info["cli_arg"]
                if cli_info["short_arg"]:
                    param["short_arg"] = cli_info["short_arg"]

            parameters.append(param)

        applications[app_id] = {
            "name": app_id,
            "description": f"{app_id} application",
            "parameters": parameters,
        }

    config = {
        "job_submit_config": {
            "version": "1.0",
            "description": "Appform job submission configuration (auto-generated from XML)",
            "applications": applications,
        }
    }

    with open(output_file, "w", encoding="utf-8") as f:
        yaml.dump(
            config,
            f,
            default_flow_style=False,
            allow_unicode=True,
            sort_keys=False,
            width=120,
        )

    print(f"  YAML: {output_file}  ({len(applications)} applications)")


# ---------------------------------------------------------------------------
# Excel output (requires pandas + openpyxl)
# ---------------------------------------------------------------------------

_COLUMNS_ORDER = [
    "appid/应用 id",
    "字段",
    "字段中文说明",
    "默认值",
    "选项值",
    "选项显示",
    "是否必填",
    "正则过滤字符",
    "源文件",
]


def _build_dataframe(all_data: List[Dict[str, str]]):
    """Build a pandas DataFrame from collected data."""
    import pandas as pd

    df = pd.DataFrame(all_data)
    for col in _COLUMNS_ORDER:
        if col not in df.columns:
            df[col] = ""
    return df[_COLUMNS_ORDER]


def create_excel_table(all_data: List[Dict[str, str]], output_file: str) -> None:
    """
    Create an Excel table with formatting.

    Requires: pandas, openpyxl (``pip install pandas openpyxl``)
    """
    if not _has_pandas() or not _has_openpyxl():
        raise ImportError(
            "pandas and openpyxl are required for Excel output. "
            "Install: pip install pandas openpyxl"
        )

    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if not all_data:
        print(f"  Warning: no data, skipping Excel: {output_file}")
        return

    df = _build_dataframe(all_data)

    wb = Workbook()
    ws = wb.active
    ws.title = "表单配置汇总"

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border

    for row_idx, row_data in enumerate(df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(
                row=row_idx, column=col_idx, value=value if pd.notna(value) else ""
            )
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            cell.border = thin_border

    # Merge appid cells
    current_app_id = None
    merge_start_row = None
    for row_idx in range(2, ws.max_row + 1):
        app_id = ws.cell(row=row_idx, column=1).value
        if app_id != current_app_id:
            if merge_start_row and row_idx - merge_start_row > 1:
                ws.merge_cells(
                    start_row=merge_start_row,
                    end_row=row_idx - 1,
                    start_column=1,
                    end_column=1,
                )
            current_app_id = app_id
            merge_start_row = row_idx
    if merge_start_row and ws.max_row - merge_start_row > 0:
        ws.merge_cells(
            start_row=merge_start_row, end_row=ws.max_row, start_column=1, end_column=1
        )

    column_widths = {
        "A": 18,
        "B": 25,
        "C": 20,
        "D": 15,
        "E": 30,
        "F": 45,
        "G": 12,
        "H": 30,
        "I": 40,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 25

    wb.save(output_file)
    print(f"  Excel: {output_file}")


# ---------------------------------------------------------------------------
# CSV output (requires pandas)
# ---------------------------------------------------------------------------


def create_csv_table(all_data: List[Dict[str, str]], output_file: str) -> None:
    """
    Create a CSV table.

    Requires: pandas (``pip install pandas``)
    """
    if not _has_pandas():
        raise ImportError(
            "pandas is required for CSV output. Install: pip install pandas"
        )

    if not all_data:
        print(f"  Warning: no data, skipping CSV: {output_file}")
        return

    df = _build_dataframe(all_data)
    df.to_csv(output_file, index=False, encoding="utf-8-sig")
    print(f"  CSV: {output_file}")


# ---------------------------------------------------------------------------
# Markdown output (stdlib only + optional pandas)
# ---------------------------------------------------------------------------


def create_markdown_table(
    all_data: List[Dict[str, str]],
    output_file: str,
    source_files: Optional[List[str]] = None,
) -> None:
    """
    Create a Markdown table. No extra dependencies required.
    """
    if not all_data:
        print(f"  Warning: no data, skipping Markdown: {output_file}")
        return

    if _has_pandas():
        import pandas as pd

        df = _build_dataframe(all_data)
        columns = list(df.columns)
        rows = [list(row) for _, row in df.iterrows()]
    else:
        columns = _COLUMNS_ORDER
        rows = []
        for row in all_data:
            rows.append([row.get(c, "") for c in columns])

    md = f"# 表单配置汇总文档\n\n"
    md += f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  \n"
    md += f"**记录数**: {len(all_data)}  \n"
    if source_files:
        md += f"**应用数**: {len(source_files)}\n\n"
        md += f"## 源文件列表\n\n"
        for src in source_files:
            md += f"- `{src}`\n"
    md += f"\n## 配置详情\n\n"

    md += "| " + " | ".join(columns) + " |\n"
    md += "| " + " | ".join(["---"] * len(columns)) + " |\n"

    for row in rows:
        row_values = []
        for value in row:
            value_str = str(value) if value is not None else ""
            value_str = value_str.replace("|", "\\|").replace("\n", "<br>")
            row_values.append(value_str)
        md += "| " + " | ".join(row_values) + " |\n"

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(md)

    print(f"  Markdown: {output_file}")


# ---------------------------------------------------------------------------
# High-level API
# ---------------------------------------------------------------------------


def check_available_formats() -> Dict[str, bool]:
    """Check which output formats are available in the current environment."""
    return {
        "yaml": _has_yaml(),
        "excel": _has_pandas() and _has_openpyxl(),
        "csv": _has_pandas(),
        "markdown": True,
    }


def export(
    xml_source: str,
    output_formats: Optional[List[str]] = None,
    output_dir: Optional[str] = None,
    output_filename: Optional[str] = None,
    recursive: bool = True,
    pattern: str = "view.xml",
) -> None:
    """
    High-level API: scan XML files and export in requested formats.

    Args:
        xml_source: Directory to scan, or path to a single XML file
        output_formats: List of formats: 'yaml', 'excel', 'csv', 'markdown'
        output_dir: Output directory (default: current directory)
        output_filename: Base filename without extension (default: all_apps_form_table)
        recursive: Search subdirectories (only when xml_source is a directory)
        pattern: Filename pattern (default: view.xml)
    """
    # Resolve available formats
    available = check_available_formats()

    if output_formats is None:
        output_formats = [f for f, ok in available.items() if ok]
    else:
        for fmt in output_formats:
            if fmt not in available:
                raise ValueError(
                    f"Unknown format: {fmt}. Supported: {', '.join(available)}"
                )
            if not available[fmt]:
                pkg = {"yaml": "PyYAML", "excel": "pandas openpyxl", "csv": "pandas"}[
                    fmt
                ]
                raise ImportError(
                    f"Format '{fmt}' requires {pkg}. Install: pip install {pkg}"
                )

    # Collect XML files
    if os.path.isfile(xml_source):
        xml_files = [xml_source]
    elif os.path.isdir(xml_source):
        xml_files = find_xml_files(xml_source, pattern, recursive)
    else:
        raise FileNotFoundError(f"Source not found: {xml_source}")

    if not xml_files:
        print(f"No XML files found in: {xml_source}")
        return

    print(f"\nFound {len(xml_files)} XML file(s)")
    print("=" * 60)

    # Parse data
    all_data, source_files = collect_xml_data(xml_files)

    if not all_data:
        print("\nNo data extracted")
        return

    print(
        f"\nCollected {len(all_data)} records from {len(source_files)} application(s)"
    )
    print("=" * 60)

    # Prepare output paths
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    base_name = output_filename or "all_apps_form_table"
    base_path = os.path.join(output_dir, base_name) if output_dir else base_name

    # Generate outputs
    generators = {
        "yaml": lambda: create_sdk_yaml(all_data, f"{base_path}.yaml"),
        "excel": lambda: create_excel_table(all_data, f"{base_path}.xlsx"),
        "csv": lambda: create_csv_table(all_data, f"{base_path}.csv"),
        "markdown": lambda: create_markdown_table(
            all_data, f"{base_path}.md", source_files
        ),
    }

    for fmt in output_formats:
        generators[fmt]()

    print(f"\nDone!")


def main() -> None:
    """CLI entry point for standalone usage."""
    import argparse

    parser = argparse.ArgumentParser(
        description="XML application form parser — export to Excel/CSV/Markdown/SDK YAML",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Available formats in current environment:
{chr(10).join(f'  {fmt:12s} {"OK" if ok else "MISSING: pip install " + {"yaml": "PyYAML", "excel": "pandas openpyxl", "csv": "pandas", "markdown": "-"}[fmt]}' for fmt, ok in check_available_formats().items())}

Examples:
  python -m appform_sdk.xml2table -a -r -d .
  python -m appform_sdk.xml2table -a -r -d . -o yaml -O ./output
  python -m appform_sdk.xml2table -f view.xml -o yaml,csv
        """,
    )

    parser.add_argument("-f", "--file", help="Single XML file path")
    parser.add_argument(
        "-a", "--all", action="store_true", help="Process all XML files in directory"
    )
    parser.add_argument(
        "-r", "--recursive", action="store_true", help="Search subdirectories"
    )
    parser.add_argument(
        "-d", "--directory", default=".", help="Search directory (default: .)"
    )
    parser.add_argument(
        "-p",
        "--pattern",
        default="view.xml",
        help="File match pattern (default: view.xml)",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Output formats, comma-separated (default: all available)",
    )
    parser.add_argument(
        "-O", "--output-dir", help="Output directory (default: current)"
    )
    parser.add_argument(
        "-n",
        "--filename",
        help="Output filename without extension (default: all_apps_form_table)",
    )

    args = parser.parse_args()

    output_formats = None
    if args.output:
        output_formats = [f.strip().lower() for f in args.output.split(",")]

    source = args.file if args.file else args.directory
    if args.file and not os.path.exists(args.file):
        print(f"File not found: {args.file}")
        return
    if not args.file and not args.all:
        parser.print_help()
        return

    export(
        xml_source=source,
        output_formats=output_formats,
        output_dir=args.output_dir,
        output_filename=args.filename,
        recursive=args.recursive if args.all else False,
        pattern=args.pattern,
    )


if __name__ == "__main__":
    main()
